# pip install selenium yaz ve paketi indir 
#selenium modülü içerisinden webdriverı ekle import et
from selenium import webdriver     #selenium internetteki işlemleri otomatize eder.
#webdriver_manager.chrome içinden  ChromeDriverManager ekle
from webdriver_manager.chrome import ChromeDriverManager
#time içinden sleep uykuyu import et ekle
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
from constants import globalConstants

class Test_DemoClass:
    #her teste çalışır örnek 3 test 3 kez çalışır
    #her testen önce çağrılır
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.Url)
        print("1")  #pyteste print işlemi çalışmaz
        self.folderPath = str(date.today()) #günün tarihini alır 
        Path(self.folderPath).mkdir(exist_ok=True) # yeni bir dosya açar



    #her testten sonra çağrılır
    def teardown_method(self):
        self.driver.quit() #tarayıcıyı kapar
        print("2")    

    def getData():  #decatorde kullanılacak kodlarda self parametresi aranmaz
        #veriyial 
        excelFile=openpyxl.load_workbook("data/invalid_login.xlsx") # excel dosyasını açar dosya yolu ile 
        selectedSheet=excelFile["Sayfa1"] # verilerin bulunduğu excel sayfası
        
        totalRows=selectedSheet.max_row #seçili sayfada maksimum satırı al 
        data=[]   # data adlı liste
        for i in range(2,totalRows+1):  #2.satırdan başlayarak maksimum satıra git
            userName = selectedSheet.cell(i,1).value #1.sütündaki değeri al 
            password = selectedSheet.cell(i,2).value #2.sütündaki değeri al
            tupleData = (userName,password) #tuple olarak tut değerleri
            data.append(tupleData) # data adlı listeye ekle
        return data    # data adlı listeyi fonksiyondan dön
        
#setup -> test_demoFunc -> teardown sırası ile çalışır
    @pytest.mark.parametrize("username,password",getData())
    def test_demoFunc(self,username,password):
        self.waitForElementVisible((By.ID,"user-name")) 
        usernameInput= self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password")) 
        passwordInput= self.driver.find_element(By.ID,"password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn= self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage=self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png") #ekran görüntüsü alır
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"   
    @pytest.mark.parametrize("username,password",getData())
    def test_demo2(self,username,password):
        
        self.waitForElementVisible((By.ID,"user-name")) 
        usernameInput= self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password")) 
        passwordInput= self.driver.find_element(By.ID,"password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn= self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage=self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png") #ekran görüntüsü alır
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"   

    @pytest.mark.parametrize("username,password",getData())
   # @pytest.mark.skip() #ön ek prefix@ ile çalışır burayı bu testi atla işlemi bitir
    def test_invalid_login(self,username,password):
        
        self.waitForElementVisible((By.ID,"user-name")) 
        usernameInput= self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password")) 
        passwordInput= self.driver.find_element(By.ID,"password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn= self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        errorMessage=self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png") #ekran görüntüsü alır
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"   

    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))
    

        